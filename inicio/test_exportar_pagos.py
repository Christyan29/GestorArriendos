from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import User
from datetime import date, timedelta
from inicio.models import PerfilUsuario, Contrato, Pago
from io import BytesIO
from openpyxl import load_workbook

class ExportarPagosTest(TestCase):
    def setUp(self):
        # Se crea un usuario de arrendatario de prueba
        self.user = User.objects.create_user(username='admin', password='intesud_123')
        PerfilUsuario.objects.create(user=self.user, tipo_usuario='arrendatario')
        self.client.force_login(self.user)

        # Se crea un contrato de prueba
        contrato = Contrato.objects.create(
            arrendatario=self.user,
            numero_departamento='C-001',
            fecha_inicio=date.today(),
            fecha_fin=date.today() + timedelta(days=30),
            tarifa_mensual=150,
            estado='activo'
        )

        # Se crea un pago del arrendatario
        pago = Pago.objects.create(
            contrato=contrato,
            periodo=date.today(),
            fecha_vencimiento=date.today() + timedelta(days=5),
            monto=100,
            fecha_pago=date.today(),
            monto_pagado=100,
            estado='pagado'
        )
        pago.save(update_fields=['estado'])

    def test_exportar_excel(self):
        url = reverse('exportar_pagos_excel')
        response = self.client.get(url)

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        self.assertEqual(ws['A2'].value, 'C-001')  # número de departamento
        self.assertEqual(ws['C2'].value, 100)      # monto pagado
        self.assertEqual(ws['D2'].value, 'pagado') # estado del pago
